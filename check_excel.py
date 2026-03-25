import openpyxl
wb = openpyxl.load_workbook('Google AD SOP 9-Mar v1.xlsx')
print('工作表列表:', wb.sheetnames)
for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    print(f'\n=== {sheet_name} ===')
    print(f'行数: {sheet.max_row}, 列数: {sheet.max_column}')
    # 打印前5行数据
    for i, row in enumerate(sheet.iter_rows(values_only=True), 1):
        if i > 5:
            break
        print(row)
