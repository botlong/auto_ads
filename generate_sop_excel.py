import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 从代码中提取的真实 SOP 数据
sop_data = [
    # Stage 1 - 基础设置 (S001-S011)
    {
        "stage": 1,
        "campaign_type": "All",
        "module": "转化追踪校验",
        "sop_id": "S001",
        "sop_name": "转化目标配置合规性诊断",
        "goal": "检查",
        "automation_level": "L2 (Diagnostic)",
        "kpi": "Primary conversions, Secondary conversions",
        "trigger": "账户级",
        "action": """1. 获取全部 Conversion Actions
2. 匹配业务目标分类
3. 校验 Primary/Secondary 设置
4. 检查重复计数问题

IF Category 为 Purchase/Sale AND Primary = Secondary THEN 标记为高风险
IF 同一页面存在多个 Primary 转化 THEN 标记为重复计数""",
        "priority": "P1",
        "frequency": "每周",
        "output": "异常报告、转化目标修改建议"
    },
    {
        "stage": 1,
        "campaign_type": "All",
        "module": "转化追踪校验",
        "sop_id": "S002",
        "sop_name": "转化价值与计数完整性监控",
        "goal": "检查",
        "automation_level": "L1 (Monitoring)",
        "kpi": "Conversion value, All conversions",
        "trigger": "账户级",
        "action": """1. 检查 Purchase 类转化价值是否为 0
2. 检查计数方式是否正确
3. 检查转化数据一致性

IF Category IN (Purchase, Booking) AND Conversion Value == 0 THEN 标记为价值缺失
IF Counting Method == Every AND Category == Page View THEN 标记为重复计数""",
        "priority": "P1",
        "frequency": "每天",
        "output": "转化价值异常报告"
    },
    {
        "stage": 1,
        "campaign_type": "All",
        "module": "转化追踪校验",
        "sop_id": "S003",
        "sop_name": "进阶转化功能优化建议",
        "goal": "优化",
        "automation_level": "L3 (Recommendation)",
        "kpi": "CVR, Conversions",
        "trigger": "账户级",
        "action": """1. 检查 Tag 状态 (Active/Inactive/Unverified)
2. 检查 Enhanced Conversions 启用状态
3. 验证离线转化导入

IF Tag Status == INACTIVE THEN 标记为追踪失效
IF Enhanced Conversions == Disabled THEN 建议启用""",
        "priority": "P2",
        "frequency": "每周",
        "output": "优化建议报告"
    },
    {
        "stage": 1,
        "campaign_type": "All",
        "module": "业务目标校准",
        "sop_id": "S004",
        "sop_name": "Campaign 业务目标一致性校正",
        "goal": "检查",
        "automation_level": "L2 (Diagnostic)",
        "kpi": "Campaign settings",
        "trigger": "Campaign 级",
        "action": """1. 获取 Campaign Bidding Strategy
2. 获取业务类型 (Lead Gen / Ecommerce)
3. 校验匹配度

IF Business Type == Lead Gen AND Bidding NOT IN (tCPA, Max Conv) THEN 标记为不匹配
IF Business Type == Ecommerce AND Bidding NOT IN (tROAS, Max Conv Value) THEN 标记为不匹配""",
        "priority": "P1",
        "frequency": "Campaign 创建时",
        "output": "出价策略调整建议"
    },
    {
        "stage": 1,
        "campaign_type": "All",
        "module": "出价策略优化",
        "sop_id": "S005",
        "sop_name": "竞价策略效能诊断与放量调整",
        "goal": "优化",
        "automation_level": "L3 (Recommendation)",
        "kpi": "CPA, ROAS, Impression Share",
        "trigger": "Campaign 级",
        "action": """1. 检查 Learning Status
2. 分析实际 vs 目标 CPA/ROAS
3. 评估展示份额损失
4. 生成调整建议

IF Actual CPA < Target * 0.7 THEN 建议收紧目标
IF Actual CPA > Target * 1.2 AND IS < 50% THEN 建议放宽目标
IF ROAS > Target * 1.2 AND Lost IS (Rank) > 30% THEN 建议提升目标 ROAS""",
        "priority": "P1",
        "frequency": "每天",
        "output": "出价策略调整建议"
    },
    {
        "stage": 1,
        "campaign_type": "All",
        "module": "预算管理",
        "sop_id": "S006",
        "sop_name": "预算瓶颈识别与效能调配",
        "goal": "优化",
        "automation_level": "L3 (Recommendation)",
        "kpi": "Budget, Lost IS (Budget), ROAS",
        "trigger": "Campaign 级",
        "action": """1. 计算预算使用率
2. 分析 Lost IS (Budget/Rank)
3. 评估 Campaign 效率
4. 生成预算调整建议

IF ROAS > Target * 1.1 AND Lost IS (Budget) > 15% THEN 建议增加预算 20%
IF ROAS < Target * 0.8 AND Budget Utilization > 80% THEN 建议减少预算 20%""",
        "priority": "P1",
        "frequency": "每天",
        "output": "预算调整建议"
    },
    {
        "stage": 1,
        "campaign_type": "Search",
        "module": "账户结构",
        "sop_id": "S007",
        "sop_name": "品牌词与非品牌词隔离",
        "goal": "检查",
        "automation_level": "L2 (Diagnostic)",
        "kpi": "Campaign structure",
        "trigger": "Campaign 级",
        "action": """1. 检查 Brand Campaign 是否包含非品牌词
2. 检查非品牌 Campaign 是否包含品牌词
3. 标记混合意图问题

IF Is Brand Campaign == True AND Has Non-brand Keywords == True THEN 标记为结构问题""",
        "priority": "P1",
        "frequency": "每周",
        "output": "Campaign 拆分建议"
    },
    {
        "stage": 1,
        "campaign_type": "Search",
        "module": "账户结构",
        "sop_id": "S008",
        "sop_name": "搜索与展示网络隔离",
        "goal": "检查",
        "automation_level": "L2 (Diagnostic)",
        "kpi": "Network type",
        "trigger": "Campaign 级",
        "action": """1. 检查 Campaign Network Type
2. 识别 Search + Display 混合

IF Network Type == SEARCH_WITH_DISPLAY THEN 标记为网络混合
建议：关闭 Display Network 或创建独立 Display Campaign""",
        "priority": "P1",
        "frequency": "Campaign 创建时",
        "output": "网络设置调整建议"
    },
    {
        "stage": 1,
        "campaign_type": "All",
        "module": "账户结构",
        "sop_id": "S009",
        "sop_name": "账户结构粒度优化",
        "goal": "优化",
        "automation_level": "L3 (Recommendation)",
        "kpi": "AdGroup count",
        "trigger": "Campaign 级",
        "action": """1. 统计 Campaign 下 AdGroup 数量
2. 检查关键词主题集中度
3. 评估是否需要拆分

IF AdGroup Count > 15 THEN 标记为粒度过粗
建议：按产品/服务/主题拆分 Campaign""",
        "priority": "P2",
        "frequency": "每月",
        "output": "Campaign 拆分建议"
    },
    {
        "stage": 1,
        "campaign_type": "Ecommerce",
        "module": "商品效率分层",
        "sop_id": "S010",
        "sop_name": "购物/PMax 商品效率分层",
        "goal": "优化",
        "automation_level": "L3 (Recommendation)",
        "kpi": "ROAS by product",
        "trigger": "Product 级",
        "action": """1. 计算各商品 ROAS
2. 对比目标 ROAS
3. 划分效率层级

IF ROAS > Target * 1.2 THEN 标记为 Hero Product (增加预算)
IF ROAS < Target * 0.7 THEN 标记为 Low Efficiency (降权或排除)""",
        "priority": "P2",
        "frequency": "每周",
        "output": "商品分层报告、预算分配建议"
    },
    {
        "stage": 1,
        "campaign_type": "All",
        "module": "多触归因",
        "sop_id": "S011",
        "sop_name": "跨渠道归因窗口一致性检查",
        "goal": "检查",
        "automation_level": "L2 (Diagnostic)",
        "kpi": "Attribution model, Conversion window",
        "trigger": "账户级",
        "action": """1. 检查归因模型设置
2. 检查转化时间窗口
3. 对比行业最佳实践

建议：使用 Data-driven 归因模型
建议：转化窗口设置为 30/90 天""",
        "priority": "P2",
        "frequency": "每月",
        "output": "归因设置建议"
    },

    # Stage 2 - 组件级优化 (S012-S053)
    {
        "stage": 2,
        "campaign_type": "Search",
        "module": "搜索词优化",
        "sop_id": "S012",
        "sop_name": "搜索词负向过滤",
        "goal": "清理",
        "automation_level": "L4 (Auto-fix)",
        "kpi": "Search term cost, conversions",
        "trigger": "搜索词级",
        "action": """1. 获取搜索词报告
2. 筛选高花费无转化搜索词
3. 生成负向词建议

IF Cost > $30 AND Conversions == 0 THEN 建议添加为精确负向词""",
        "priority": "P1",
        "frequency": "每天",
        "output": "负向词列表"
    },
    {
        "stage": 2,
        "campaign_type": "Search",
        "module": "搜索词优化",
        "sop_id": "S013",
        "sop_name": "高意图搜索词拓新",
        "goal": "拓展",
        "automation_level": "L3 (Recommendation)",
        "kpi": "Search term CVR",
        "trigger": "搜索词级",
        "action": """1. 分析高转化搜索词
2. 识别新意图模式
3. 生成关键词建议

IF CVR > Campaign Avg * 1.5 THEN 建议添加为新关键词""",
        "priority": "P2",
        "frequency": "每周",
        "output": "新关键词建议"
    },
    {
        "stage": 2,
        "campaign_type": "Search",
        "module": "搜索词优化",
        "sop_id": "S014",
        "sop_name": "匹配类型流量健康度检查",
        "goal": "检查",
        "automation_level": "L2 (Diagnostic)",
        "kpi": "Match type distribution",
        "trigger": "Campaign 级",
        "action": """1. 分析各匹配类型表现
2. 检查流量分布
3. 识别优化机会

建议：Phrase Match 占比 40-60%
建议：Exact Match 占比 20-40%
建议：Broad Match 占比 < 20%""",
        "priority": "P2",
        "frequency": "每月",
        "output": "匹配类型优化建议"
    },
    {
        "stage": 2,
        "campaign_type": "Search",
        "module": "关键词管理",
        "sop_id": "S015",
        "sop_name": "低效关键词清理",
        "goal": "清理",
        "automation_level": "L4 (Auto-fix)",
        "kpi": "Keyword CPA, ROAS",
        "trigger": "关键词级",
        "action": """1. 计算各关键词效率
2. 对比目标 CPA/ROAS
3. 生成清理建议

IF CPA > Target * 1.5 AND Conversions < 3 THEN 建议暂停
IF ROAS < Target * 0.7 AND Cost > $50 THEN 建议暂停""",
        "priority": "P1",
        "frequency": "每周",
        "output": "关键词暂停建议"
    },
    {
        "stage": 2,
        "campaign_type": "Search",
        "module": "关键词管理",
        "sop_id": "S016",
        "sop_name": "关键词质量得分诊断",
        "goal": "优化",
        "automation_level": "L3 (Recommendation)",
        "kpi": "Quality Score",
        "trigger": "关键词级",
        "action": """1. 获取 Quality Score 数据
2. 分析 Ad Relevance, LP Experience, Expected CTR
3. 生成优化建议

IF Quality Score < 5 THEN 诊断具体维度
IF Ad Relevance == Below Average THEN 优化广告文案""",
        "priority": "P2",
        "frequency": "每周",
        "output": "质量得分优化建议"
    },
    {
        "stage": 2,
        "campaign_type": "Search",
        "module": "关键词管理",
        "sop_id": "S017",
        "sop_name": "出价水位动态调整",
        "goal": "优化",
        "automation_level": "L4 (Auto-fix)",
        "kpi": "CPC, First page CPC, Top of page CPC",
        "trigger": "关键词级",
        "action": """1. 对比当前出价与页面位置出价
2. 分析 Impression Share 损失
3. 调整出价

IF Lost IS (Rank) > 30% AND CPC < Top of Page CPC THEN 建议提升出价 10%""",
        "priority": "P2",
        "frequency": "每天",
        "output": "出价调整建议"
    },
    {
        "stage": 2,
        "campaign_type": "All",
        "module": "广告文案优化",
        "sop_id": "S018",
        "sop_name": "Responsive Search Ad 强度补全",
        "goal": "优化",
        "automation_level": "L3 (Recommendation)",
        "kpi": "Ad Strength",
        "trigger": "广告级",
        "action": """1. 检查 Ad Strength Rating
2. 统计 Headlines/Descriptions 数量
3. 检查 Pin 设置

IF Headlines < 10 THEN 建议添加标题
IF Descriptions < 3 THEN 建议添加描述
IF Pinned Count > 3 THEN 建议减少固定""",
        "priority": "P2",
        "frequency": "每周",
        "output": "广告优化建议"
    },
    {
        "stage": 2,
        "campaign_type": "All",
        "module": "广告文案优化",
        "sop_id": "S019",
        "sop_name": "低 CTR 素材清理",
        "goal": "清理",
        "automation_level": "L4 (Auto-fix)",
        "kpi": "Asset CTR",
        "trigger": "素材级",
        "action": """1. 分析各素材 CTR
2. 识别低表现素材
3. 生成替换建议

IF CTR < 1% AND Impressions > 1000 THEN 建议替换素材""",
        "priority": "P2",
        "frequency": "每周",
        "output": "素材替换建议"
    },
    {
        "stage": 2,
        "campaign_type": "All",
        "module": "广告文案优化",
        "sop_id": "S020",
        "sop_name": "Offer 与 CTA 注入",
        "goal": "优化",
        "automation_level": "L3 (Recommendation)",
        "kpi": "CVR",
        "trigger": "广告级",
        "action": """1. 检查广告文案中 Offer 信息
2. 检查 CTA 清晰度
3. 对比竞品文案

建议：标题包含具体优惠 (如 'Save 20%')
建议：描述包含明确 CTA (如 'Shop Now')""",
        "priority": "P2",
        "frequency": "每月",
        "output": "文案优化建议"
    },
    {
        "stage": 2,
        "campaign_type": "All",
        "module": "Assets 管理",
        "sop_id": "S021",
        "sop_name": "Assets 使用率与饱和度",
        "goal": "检查",
        "automation_level": "L2 (Diagnostic)",
        "kpi": "Assets count",
        "trigger": "Campaign 级",
        "action": """1. 统计各类型 Assets 数量
2. 检查饱和度
3. 识别缺失类型

建议：Sitelinks >= 4
建议：Callouts >= 6
建议：Images >= 3""",
        "priority": "P2",
        "frequency": "每月",
        "output": "Assets 补充建议"
    },
    {
        "stage": 2,
        "campaign_type": "All",
        "module": "Assets 管理",
        "sop_id": "S022",
        "sop_name": "季节性/促销 Sitelink 更新",
        "goal": "优化",
        "automation_level": "L3 (Recommendation)",
        "kpi": "Sitelink CTR",
        "trigger": "时间触发",
        "action": """1. 检查当前 Sitelink 时效性
2. 对比促销日历
3. 生成更新建议

IF Days Since Updated > 30 THEN 建议刷新 Sitelink
IF 促销活动期间 THEN 添加促销 Sitelink""",
        "priority": "P2",
        "frequency": "促销期间",
        "output": "Sitelink 更新建议"
    },
    {
        "stage": 2,
        "campaign_type": "All",
        "module": "Assets 管理",
        "sop_id": "S023",
        "sop_name": "图片与视频素材合规性检查",
        "goal": "检查",
        "automation_level": "L2 (Diagnostic)",
        "kpi": "Asset approval status",
        "trigger": "素材级",
        "action": """1. 检查素材审核状态
2. 识别拒登素材
3. 生成替换建议

IF Status == DISAPPROVED THEN 标记为拒登并建议替换""",
        "priority": "P0",
        "frequency": "每天",
        "output": "素材合规报告"
    },
    {
        "stage": 2,
        "campaign_type": "Search",
        "module": "落地页体验",
        "sop_id": "S024",
        "sop_name": "移动与桌面端 LP 速度优化",
        "goal": "优化",
        "automation_level": "L3 (Recommendation)",
        "kpi": "Landing page speed",
        "trigger": "页面级",
        "action": """1. 检测落地页加载速度
2. 对比基准值
3. 生成优化建议

IF Load Time > 3s THEN 标记为加载缓慢
建议：压缩图片、启用 CDN、减少重定向""",
        "priority": "P1",
        "frequency": "每周",
        "output": "落地页速度优化建议"
    },
    {
        "stage": 2,
        "campaign_type": "Search",
        "module": "落地页体验",
        "sop_id": "S025",
        "sop_name": "LP 内容与广告一致性检查",
        "goal": "检查",
        "automation_level": "L2 (Diagnostic)",
        "kpi": "Bounce rate",
        "trigger": "页面级",
        "action": """1. 对比广告文案与落地页内容
2. 分析跳出率
3. 识别不一致问题

IF Bounce Rate > 70% THEN 检查内容一致性
建议：落地页标题与广告标题保持一致""",
        "priority": "P1",
        "frequency": "每月",
        "output": "一致性检查报告"
    },
    {
        "stage": 2,
        "campaign_type": "Search",
        "module": "地域优化",
        "sop_id": "S026",
        "sop_name": "地域表现分析与排除",
        "goal": "优化",
        "automation_level": "L4 (Auto-fix)",
        "kpi": "CPA by location",
        "trigger": "地域级",
        "action": """1. 计算各地域 CPA/ROAS
2. 对比平均值
3. 生成排除建议

IF CPA > Avg * 1.5 AND Cost > $50 THEN 建议排除该地域""",
        "priority": "P2",
        "frequency": "每周",
        "output": "地域排除建议"
    },
    {
        "stage": 2,
        "campaign_type": "Search",
        "module": "地域优化",
        "sop_id": "S027",
        "sop_name": "高表现地域出价提升",
        "goal": "优化",
        "automation_level": "L4 (Auto-fix)",
        "kpi": "ROAS by location",
        "trigger": "地域级",
        "action": """1. 识别高 ROAS 地域
2. 计算出价调整幅度
3. 应用 Bid Modifier

IF ROAS > Avg * 1.2 AND Conversions >= 3 THEN 建议提升出价 15%""",
        "priority": "P2",
        "frequency": "每周",
        "output": "出价调整建议"
    },
    {
        "stage": 2,
        "campaign_type": "Search",
        "module": "地域优化",
        "sop_id": "S028",
        "sop_name": "Proximity Targeting 调整",
        "goal": "优化",
        "automation_level": "L3 (Recommendation)",
        "kpi": "Store visits",
        "trigger": "本地业务",
        "action": """1. 分析门店周边表现
2. 调整半径设置
3. 优化本地出价

建议：根据门店密度调整半径 (1-10 miles)
建议：高峰时段提升本地出价""",
        "priority": "P2",
        "frequency": "每月",
        "output": "本地投放优化建议"
    },
    {
        "stage": 2,
        "campaign_type": "Search",
        "module": "受众优化",
        "sop_id": "S029",
        "sop_name": "第一方受众分层与定向",
        "goal": "优化",
        "automation_level": "L3 (Recommendation)",
        "kpi": "Audience list size, CTR",
        "trigger": "受众级",
        "action": """1. 分析第一方受众表现
2. 分层 (高价值/中价值/低价值)
3. 生成定向策略

建议：对高价值受众提升出价 20%
建议：对流失受众开启再营销""",
        "priority": "P2",
        "frequency": "每月",
        "output": "受众分层策略"
    },
    {
        "stage": 2,
        "campaign_type": "Search",
        "module": "受众优化",
        "sop_id": "S030",
        "sop_name": "相似受众与自定义受众扩展",
        "goal": "拓展",
        "automation_level": "L3 (Recommendation)",
        "kpi": "Lookalike performance",
        "trigger": "受众级",
        "action": """1. 创建高价值客户相似受众
2. 测试不同相似度 (1%-10%)
3. 监控扩展受众表现

建议：从 1% 相似度开始测试
建议：排除已有客户避免重复触达""",
        "priority": "P2",
        "frequency": "每季度",
        "output": "受众扩展建议"
    },
    {
        "stage": 2,
        "campaign_type": "Search",
        "module": "受众优化",
        "sop_id": "S031",
        "sop_name": "再营销受众排除与观察",
        "goal": "优化",
        "automation_level": "L4 (Auto-fix)",
        "kpi": "Audience overlap",
        "trigger": "受众级",
        "action": """1. 检查受众重叠度
2. 识别重复定向
3. 优化排除设置

建议：已转化用户排除 30 天
建议：再营销 Campaign 与搜索 Campaign 排除重叠""",
        "priority": "P2",
        "frequency": "每月",
        "output": "受众排除优化"
    },
    {
        "stage": 2,
        "campaign_type": "Search",
        "module": "受众优化",
        "sop_id": "S032",
        "sop_name": "受众表现分析与排除",
        "goal": "清理",
        "automation_level": "L4 (Auto-fix)",
        "kpi": "CPA by audience",
        "trigger": "受众级",
        "action": """1. 计算各受众 CPA/ROAS
2. 识别低效受众
3. 生成排除建议

IF CPA > Target * 1.5 AND Cost > $50 THEN 建议排除该受众""",
        "priority": "P2",
        "frequency": "每周",
        "output": "受众排除建议"
    },
    {
        "stage": 2,
        "campaign_type": "Search",
        "module": "设备优化",
        "sop_id": "S033",
        "sop_name": "设备表现差异分析",
        "goal": "检查",
        "automation_level": "L2 (Diagnostic)",
        "kpi": "CVR by device",
        "trigger": "设备级",
        "action": """1. 计算各设备转化率
2. 对比移动端与桌面端
3. 识别显著差异

IF Mobile CVR < Desktop CVR * 0.5 THEN 标记为移动端体验问题""",
        "priority": "P1",
        "frequency": "每周",
        "output": "设备表现报告"
    },
    {
        "stage": 2,
        "campaign_type": "Search",
        "module": "设备优化",
        "sop_id": "S034",
        "sop_name": "设备出价调整",
        "goal": "优化",
        "automation_level": "L4 (Auto-fix)",
        "kpi": "CPA by device",
        "trigger": "设备级",
        "action": """1. 计算设备 CPA 差异
2. 对比平均值
3. 调整 Bid Modifier

IF Device CPA > Avg * 1.3 THEN 降低出价 15%
IF Device CPA < Avg * 0.7 THEN 提高出价 15%""",
        "priority": "P2",
        "frequency": "每周",
        "output": "设备出价调整"
    },
    {
        "stage": 2,
        "campaign_type": "Search",
        "module": "时段优化",
        "sop_id": "S035",
        "sop_name": "时段表现分析与出价调整",
        "goal": "优化",
        "automation_level": "L4 (Auto-fix)",
        "kpi": "CPA by hour",
        "trigger": "时段级",
        "action": """1. 计算各时段 CPA
2. 识别高效/低效时段
3. 应用 Ad Schedule Bid Adjustment

IF Hour CPA < Avg * 0.7 THEN 提升出价 15%
IF Hour CPA > Avg * 1.3 THEN 降低出价 20%""",
        "priority": "P2",
        "frequency": "每周",
        "output": "时段出价调整"
    },
    {
        "stage": 2,
        "campaign_type": "Search",
        "module": "时段优化",
        "sop_id": "S036",
        "sop_name": "Ad Schedule 饱和检查",
        "goal": "检查",
        "automation_level": "L2 (Diagnostic)",
        "kpi": "Schedule coverage",
        "trigger": "Campaign 级",
        "action": """1. 检查 Ad Schedule 覆盖度
2. 识别投放空白时段
3. 评估预算分配

建议：工作日 8:00-22:00 覆盖
建议：周末根据业务调整""",
        "priority": "P2",
        "frequency": "每月",
        "output": "时段覆盖报告"
    },
    {
        "stage": 2,
        "campaign_type": "Search",
        "module": "竞品分析",
        "sop_id": "S037",
        "sop_name": "竞品拍卖洞察分析",
        "goal": "分析",
        "automation_level": "L2 (Diagnostic)",
        "kpi": "Auction insights",
        "trigger": "Campaign 级",
        "action": """1. 获取 Auction Insights 数据
2. 分析竞品展示份额
3. 评估排名竞争力

IF Competitor IS > Your IS + 20% THEN 标记为竞争劣势
建议：提高出价或优化质量得分""",
        "priority": "P2",
        "frequency": "每周",
        "output": "竞品分析报告"
    },
    {
        "stage": 2,
        "campaign_type": "All",
        "module": "政策合规",
        "sop_id": "S038",
        "sop_name": "广告政策合规性检查",
        "goal": "检查",
        "automation_level": "L1 (Monitoring)",
        "kpi": "Approval status",
        "trigger": "广告级",
        "action": """1. 检查广告审核状态
2. 识别拒登原因
3. 生成修复建议

IF Status == DISAPPROVED THEN 立即标记 P0 告警
建议：根据拒登原因修改文案或素材""",
        "priority": "P0",
        "frequency": "每天",
        "output": "合规检查报告"
    },
    {
        "stage": 2,
        "campaign_type": "Shopping/PMax",
        "module": "Feed 优化",
        "sop_id": "S039",
        "sop_name": "商品 Feed 字段完整性检查",
        "goal": "检查",
        "automation_level": "L2 (Diagnostic)",
        "kpi": "Feed attribute coverage",
        "trigger": "商品级",
        "action": """1. 检查 Feed 必填字段
2. 识别缺失属性
3. 生成补充建议

建议：Title 包含品牌+产品类型+关键属性
建议：完整填写 size, color, material 等属性""",
        "priority": "P1",
        "frequency": "每周",
        "output": "Feed 完整性报告"
    },
    {
        "stage": 2,
        "campaign_type": "Shopping/PMax",
        "module": "Feed 优化",
        "sop_id": "S040",
        "sop_name": "商品标题与描述优化",
        "goal": "优化",
        "automation_level": "L3 (Recommendation)",
        "kpi": "Title quality score",
        "trigger": "商品级",
        "action": """1. 分析标题结构
2. 检查关键词覆盖
3. 优化描述质量

建议：标题 60-150 字符
建议：前 30 字符包含核心关键词
建议：描述突出卖点和差异化""",
        "priority": "P2",
        "frequency": "每月",
        "output": "标题优化建议"
    },
    {
        "stage": 2,
        "campaign_type": "Shopping/PMax",
        "module": "Feed 优化",
        "sop_id": "S041",
        "sop_name": "高价值商品识别与放量",
        "goal": "优化",
        "automation_level": "L3 (Recommendation)",
        "kpi": "Product ROAS",
        "trigger": "商品级",
        "action": """1. 计算各商品 ROAS
2. 识别高价值商品
3. 生成放量建议

IF Product ROAS > Campaign Avg * 1.3 THEN 标记为 Hero Product
建议：创建独立 Campaign 增加预算""",
        "priority": "P1",
        "frequency": "每周",
        "output": "高价值商品报告"
    },
    {
        "stage": 2,
        "campaign_type": "Shopping/PMax",
        "module": "PMax 专项",
        "sop_id": "S042",
        "sop_name": "PMax 搜索词负向过滤",
        "goal": "清理",
        "automation_level": "L4 (Auto-fix)",
        "kpi": "Search term cost",
        "trigger": "搜索词级",
        "action": """1. 获取 PMax 搜索词报告
2. 识别高花费无转化搜索词
3. 添加账户级负向词

IF Cost > $30 AND Conversions == 0 THEN 添加为负向词""",
        "priority": "P1",
        "frequency": "每周",
        "output": "负向词列表"
    },
    {
        "stage": 2,
        "campaign_type": "Shopping/PMax",
        "module": "PMax 专项",
        "sop_id": "S043",
        "sop_name": "PMax 与搜索 Campaign 重叠分析",
        "goal": "分析",
        "automation_level": "L2 (Diagnostic)",
        "kpi": "Overlap rate",
        "trigger": "Campaign 级",
        "action": """1. 分析 PMax 与 Search 重叠
2. 计算 cannibalization rate
3. 生成优化建议

IF Overlap Rate > 30% THEN 标记为高重叠
建议：添加负向词或调整受众信号""",
        "priority": "P2",
        "frequency": "每月",
        "output": "重叠分析报告"
    },
    {
        "stage": 2,
        "campaign_type": "All",
        "module": "品牌保护",
        "sop_id": "S044",
        "sop_name": "品牌词保护 Campaign 维护",
        "goal": "检查",
        "automation_level": "L2 (Diagnostic)",
        "kpi": "Brand impression share",
        "trigger": "品牌 Campaign",
        "action": """1. 检查品牌 Campaign 展示份额
2. 监控竞品投放品牌词
3. 确保品牌词覆盖

IF Brand IS < 90% THEN 建议提升出价
建议：品牌词使用精确匹配""",
        "priority": "P1",
        "frequency": "每周",
        "output": "品牌保护报告"
    },
    {
        "stage": 2,
        "campaign_type": "All",
        "module": "竞价保护",
        "sop_id": "S045",
        "sop_name": "竞品词投放与出价策略",
        "goal": "优化",
        "automation_level": "L3 (Recommendation)",
        "kpi": "Competitor term CPA",
        "trigger": "竞品词 Campaign",
        "action": """1. 分析竞品词表现
2. 对比自有品牌词 CPA
3. 优化出价策略

建议：竞品词 CPA 控制在品牌词 1.5 倍以内
建议：使用动态搜索广告覆盖竞品词""",
        "priority": "P2",
        "frequency": "每月",
        "output": "竞品词优化建议"
    },
    {
        "stage": 2,
        "campaign_type": "Display/Video",
        "module": "展示优化",
        "sop_id": "S046",
        "sop_name": "Placement 表现分析与排除",
        "goal": "清理",
        "automation_level": "L4 (Auto-fix)",
        "kpi": "CPA by placement",
        "trigger": "Placement 级",
        "action": """1. 分析各 Placement CPA
2. 识别低效 Placement
3. 生成排除建议

IF CPA > Target * 2 AND Cost > $30 THEN 建议排除""",
        "priority": "P2",
        "frequency": "每周",
        "output": "Placement 排除列表"
    },
    {
        "stage": 2,
        "campaign_type": "Display/Video",
        "module": "展示优化",
        "sop_id": "S047",
        "sop_name": "受众信号与再营销更新",
        "goal": "优化",
        "automation_level": "L3 (Recommendation)",
        "kpi": "Audience list freshness",
        "trigger": "受众级",
        "action": """1. 检查受众列表时效性
2. 分析再营销表现
3. 更新受众策略

建议：再营销列表 30 天内更新
建议：分层再营销 (7天/30天/90天)""",
        "priority": "P2",
        "frequency": "每月",
        "output": "受众更新建议"
    },
    {
        "stage": 2,
        "campaign_type": "Display/Video",
        "module": "视频优化",
        "sop_id": "S048",
        "sop_name": "视频素材完播率优化",
        "goal": "优化",
        "automation_level": "L3 (Recommendation)",
        "kpi": "Video completion rate",
        "trigger": "素材级",
        "action": """1. 分析视频完播率
2. 识别流失点
3. 生成优化建议

IF Completion Rate < 25% THEN 标记为需优化
建议：前 5 秒抓住注意力
建议：视频时长 15-30 秒最佳""",
        "priority": "P2",
        "frequency": "每月",
        "output": "视频优化建议"
    },
    {
        "stage": 2,
        "campaign_type": "Display/Video",
        "module": "视频优化",
        "sop_id": "S049",
        "sop_name": "YouTube 受众定位优化",
        "goal": "优化",
        "automation_level": "L3 (Recommendation)",
        "kpi": "CPV, CTR",
        "trigger": "Campaign 级",
        "action": """1. 分析 YouTube 受众表现
2. 对比不同定位方式
3. 优化受众策略

建议：自定义意向受众优于兴趣受众
建议：结合再营销提升转化""",
        "priority": "P2",
        "frequency": "每月",
        "output": "YouTube 优化建议"
    },
    {
        "stage": 2,
        "campaign_type": "App",
        "module": "应用推广",
        "sop_id": "S050",
        "sop_name": "App Campaign 素材轮换",
        "goal": "优化",
        "automation_level": "L3 (Recommendation)",
        "kpi": "CPI, Install rate",
        "trigger": "素材级",
        "action": """1. 分析各素材安装率
2. 识别疲劳素材
3. 生成轮换建议

IF Install Rate < 1% AND Impressions > 5000 THEN 建议替换素材""",
        "priority": "P2",
        "frequency": "每月",
        "output": "素材轮换建议"
    },
    {
        "stage": 2,
        "campaign_type": "App",
        "module": "应用推广",
        "sop_id": "S051",
        "sop_name": "应用内事件优化目标调整",
        "goal": "优化",
        "automation_level": "L3 (Recommendation)",
        "kpi": "In-app events",
        "trigger": "应用级",
        "action": """1. 分析应用内转化漏斗
2. 识别高价值事件
3. 优化出价目标

建议：从 Install 优化转为 In-app Purchase 优化
建议：设置有价值的应用内事件""",
        "priority": "P1",
        "frequency": "每季度",
        "output": "优化目标调整建议"
    },
    {
        "stage": 2,
        "campaign_type": "All",
        "module": "智能出价",
        "sop_id": "S052",
        "sop_name": "季节性出价调整",
        "goal": "优化",
        "automation_level": "L3 (Recommendation)",
        "kpi": "Seasonal performance",
        "trigger": "时间触发",
        "action": """1. 分析历史季节性数据
2. 预测高峰期需求
3. 调整出价策略

建议：旺季前 2 周提升出价 20%
建议：淡季期间降低出价保持 ROI""",
        "priority": "P2",
        "frequency": "季节性",
        "output": "季节性出价建议"
    },
    {
        "stage": 2,
        "campaign_type": "All",
        "module": "智能出价",
        "sop_id": "S053",
        "sop_name": "新 Campaign 学习期监控",
        "goal": "检查",
        "automation_level": "L2 (Diagnostic)",
        "kpi": "Learning status",
        "trigger": "Campaign 级",
        "action": """1. 监控 Learning Status
2. 评估学习期表现
3. 避免过早干预

IF Learning Status == LEARNING THEN 建议不调整
IF Learning Status == LIMITED AND Days > 14 THEN 检查设置问题""",
        "priority": "P2",
        "frequency": "每天",
        "output": "学习期状态报告"
    },

    # Stage 3 - Workstream 执行 (S054-S074)
    {
        "stage": 3,
        "campaign_type": "All",
        "module": "先修追踪",
        "sop_id": "S054",
        "sop_name": "转化目标配置审计与修正",
        "goal": "修复",
        "automation_level": "L4 (Auto-fix)",
        "kpi": "Conversion action health",
        "trigger": "账户级",
        "action": """1. 审计所有 Conversion Actions
2. 修正 Primary/Secondary 设置
3. 确保目标与业务一致

IF 无有效 Primary 转化 THEN 设置 Purchase 为 Primary""",
        "priority": "P0",
        "frequency": "发现时立即",
        "output": "转化目标修正"
    },
    {
        "stage": 3,
        "campaign_type": "All",
        "module": "先修追踪",
        "sop_id": "S055",
        "sop_name": "转化价值追踪验证",
        "goal": "修复",
        "automation_level": "L2 (Diagnostic)",
        "kpi": "Conversion value accuracy",
        "trigger": "账户级",
        "action": """1. 验证转化价值准确性
2. 检查动态价值传输
3. 测试转化追踪

IF 转化价值异常 THEN 重新配置转化代码""",
        "priority": "P1",
        "frequency": "发现时立即",
        "output": "追踪验证报告"
    },
    {
        "stage": 3,
        "campaign_type": "Search",
        "module": "控制浪费",
        "sop_id": "S056",
        "sop_name": "高花费无转化搜索词清理",
        "goal": "清理",
        "automation_level": "L4 (Auto-fix)",
        "kpi": "Wasted spend",
        "trigger": "搜索词级",
        "action": """1. 识别高花费无转化搜索词
2. 自动添加为负向词
3. 计算节省预算

IF Cost > Target CPA * 1.5 AND Conversions == 0 THEN 自动添加为负向词""",
        "priority": "P1",
        "frequency": "每天",
        "output": "负向词添加确认"
    },
    {
        "stage": 3,
        "campaign_type": "Search",
        "module": "控制浪费",
        "sop_id": "S057",
        "sop_name": "低效地域排除",
        "goal": "清理",
        "automation_level": "L4 (Auto-fix)",
        "kpi": "CPA by location",
        "trigger": "地域级",
        "action": """1. 识别低效地域
2. 自动排除或降低出价
3. 监控效果变化

IF CPA > Avg * 2 AND Cost > Target CPA * 2 THEN 排除该地域""",
        "priority": "P1",
        "frequency": "每周",
        "output": "地域排除确认"
    },
    {
        "stage": 3,
        "campaign_type": "Search",
        "module": "控制浪费",
        "sop_id": "S058",
        "sop_name": "低效受众排除",
        "goal": "清理",
        "automation_level": "L4 (Auto-fix)",
        "kpi": "CPA by audience",
        "trigger": "受众级",
        "action": """1. 识别低效受众
2. 从 Campaign 中排除
3. 优化受众定向

IF CPA > Target * 1.5 AND Cost > $50 THEN 排除该受众""",
        "priority": "P2",
        "frequency": "每周",
        "output": "受众排除确认"
    },
    {
        "stage": 3,
        "campaign_type": "Search",
        "module": "控制浪费",
        "sop_id": "S059",
        "sop_name": "设备出价负向调整",
        "goal": "优化",
        "automation_level": "L4 (Auto-fix)",
        "kpi": "CPA by device",
        "trigger": "设备级",
        "action": """1. 分析设备表现差异
2. 自动调整 Bid Modifier
3. 减少低效设备花费

IF Device CPA > Avg * 2 THEN 降低出价 30%""",
        "priority": "P2",
        "frequency": "每周",
        "output": "设备出价调整"
    },
    {
        "stage": 3,
        "campaign_type": "All",
        "module": "预算重分配",
        "sop_id": "S060",
        "sop_name": "Campaign 间预算动态调配",
        "goal": "优化",
        "automation_level": "L4 (Auto-fix)",
        "kpi": "Cross-campaign ROAS",
        "trigger": "账户级",
        "action": """1. 分析各 Campaign 效率
2. 识别高效/低效 Campaign
3. 自动转移预算

IF Campaign ROAS > Target * 1.2 AND Lost IS > 15% THEN 从低效 Campaign 转移 20% 预算""",
        "priority": "P1",
        "frequency": "每周",
        "output": "预算转移确认"
    },
    {
        "stage": 3,
        "campaign_type": "All",
        "module": "预算重分配",
        "sop_id": "S061",
        "sop_name": "高效 Campaign 预算扩容",
        "goal": "优化",
        "automation_level": "L3 (Recommendation)",
        "kpi": "Campaign ROAS, Budget utilization",
        "trigger": "Campaign 级",
        "action": """1. 识别 ROAS 超额完成的 Campaign
2. 评估扩容潜力
3. 建议增加预算

建议：高效 Campaign 每周可增加 10-20% 预算""",
        "priority": "P1",
        "frequency": "每周",
        "output": "预算扩容建议"
    },
    {
        "stage": 3,
        "campaign_type": "All",
        "module": "预算重分配",
        "sop_id": "S062",
        "sop_name": "低效 Campaign 预算缩减",
        "goal": "优化",
        "automation_level": "L4 (Auto-fix)",
        "kpi": "Campaign ROAS",
        "trigger": "Campaign 级",
        "action": """1. 识别持续低效的 Campaign
2. 自动减少预算
3. 转移预算至高效 Campaign

IF ROAS < Target * 0.7 FOR 7 days THEN 减少预算 20%""",
        "priority": "P1",
        "frequency": "每周",
        "output": "预算缩减确认"
    },
    {
        "stage": 3,
        "campaign_type": "Search",
        "module": "优化广告",
        "sop_id": "S063",
        "sop_name": "RSA 强度补全与优化",
        "goal": "优化",
        "automation_level": "L3 (Recommendation)",
        "kpi": "Ad Strength",
        "trigger": "广告级",
        "action": """1. 检查 RSA 完整性
2. 补充缺失素材
3. 优化 Ad Strength

IF Ad Strength < Good THEN 补充标题或描述""",
        "priority": "P2",
        "frequency": "每周",
        "output": "RSA 优化建议"
    },
    {
        "stage": 3,
        "campaign_type": "Search",
        "module": "优化广告",
        "sop_id": "S064",
        "sop_name": "低表现素材替换",
        "goal": "优化",
        "automation_level": "L4 (Auto-fix)",
        "kpi": "Asset CTR",
        "trigger": "素材级",
        "action": """1. 识别低表现素材
2. 自动生成替换建议
3. 更新广告素材

IF Asset CTR < 1% AND Impressions > 1000 THEN 标记为替换""",
        "priority": "P2",
        "frequency": "每周",
        "output": "素材替换确认"
    },
    {
        "stage": 3,
        "campaign_type": "Search",
        "module": "优化广告",
        "sop_id": "S065",
        "sop_name": "促销 Offer 注入",
        "goal": "优化",
        "automation_level": "L3 (Recommendation)",
        "kpi": "CVR",
        "trigger": "时间触发",
        "action": """1. 检查促销日历
2. 更新广告 Offer
3. 添加促销素材

建议：促销期间标题包含折扣信息
建议：描述添加限时优惠 CTA""",
        "priority": "P2",
        "frequency": "促销期间",
        "output": "促销素材更新"
    },
    {
        "stage": 3,
        "campaign_type": "Search",
        "module": "优化落地页",
        "sop_id": "S066",
        "sop_name": "落地页 CVR 诊断与优化",
        "goal": "优化",
        "automation_level": "L3 (Recommendation)",
        "kpi": "LP CVR, Bounce rate",
        "trigger": "页面级",
        "action": """1. 诊断落地页问题
2. 分析 CVR 瓶颈
3. 生成优化建议

IF Bounce Rate > 70% THEN 建议优化页面速度
IF CVR < 1% THEN 建议优化表单/CTA""",
        "priority": "P1",
        "frequency": "每月",
        "output": "落地页优化建议"
    },
    {
        "stage": 3,
        "campaign_type": "Search",
        "module": "精细化定向",
        "sop_id": "S067",
        "sop_name": "设备多维出价调整",
        "goal": "优化",
        "automation_level": "L4 (Auto-fix)",
        "kpi": "CPA by device",
        "trigger": "设备级",
        "action": """1. 分析设备表现
2. 自动调整 Bid Modifier
3. 优化设备预算分配

高效设备: +15% 出价
低效设备: -20% 出价""",
        "priority": "P2",
        "frequency": "每周",
        "output": "设备出价调整"
    },
    {
        "stage": 3,
        "campaign_type": "Search",
        "module": "精细化定向",
        "sop_id": "S068",
        "sop_name": "地域多维出价调整",
        "goal": "优化",
        "automation_level": "L4 (Auto-fix)",
        "kpi": "ROAS by location",
        "trigger": "地域级",
        "action": """1. 分析地域表现
2. 调整地域出价
3. 优化地理覆盖

高 ROAS 地域: +15% 出价
低 CPA 地域: +10% 出价""",
        "priority": "P2",
        "frequency": "每周",
        "output": "地域出价调整"
    },
    {
        "stage": 3,
        "campaign_type": "Search",
        "module": "精细化定向",
        "sop_id": "S069",
        "sop_name": "时段多维出价调整",
        "goal": "优化",
        "automation_level": "L4 (Auto-fix)",
        "kpi": "CPA by hour",
        "trigger": "时段级",
        "action": """1. 分析时段表现
2. 调整 Ad Schedule
3. 优化时段出价

高效时段: +15% 出价
低效时段: -20% 出价""",
        "priority": "P2",
        "frequency": "每周",
        "output": "时段出价调整"
    },
    {
        "stage": 3,
        "campaign_type": "Search",
        "module": "精细化定向",
        "sop_id": "S070",
        "sop_name": "受众信号优化",
        "goal": "优化",
        "automation_level": "L3 (Recommendation)",
        "kpi": "Audience performance",
        "trigger": "受众级",
        "action": """1. 分析受众表现
2. 优化受众分层
3. 调整出价策略

建议：高价值受众 +20% 出价
建议：新客获取受众单独 Campaign""",
        "priority": "P2",
        "frequency": "每月",
        "output": "受众优化建议"
    },
    {
        "stage": 3,
        "campaign_type": "Shopping/PMax",
        "module": "Feed 优化",
        "sop_id": "S071",
        "sop_name": "高价值商品 Feed 优化",
        "goal": "优化",
        "automation_level": "L3 (Recommendation)",
        "kpi": "Product ROAS",
        "trigger": "商品级",
        "action": """1. 识别高价值商品
2. 优化 Feed 信息
3. 提升商品竞争力

建议：完善产品标题和描述
建议：添加促销价格""",
        "priority": "P1",
        "frequency": "每周",
        "output": "Feed 优化建议"
    },
    {
        "stage": 3,
        "campaign_type": "Shopping/PMax",
        "module": "Feed 优化",
        "sop_id": "S072",
        "sop_name": "低价值商品降权或排除",
        "goal": "清理",
        "automation_level": "L4 (Auto-fix)",
        "kpi": "Product ROAS",
        "trigger": "商品级",
        "action": """1. 识别低价值商品
2. 降低出价或排除
3. 优化商品组合

IF Product ROAS < Target * 0.5 THEN 降低出价 50% 或排除""",
        "priority": "P1",
        "frequency": "每周",
        "output": "商品降权确认"
    },
    {
        "stage": 3,
        "campaign_type": "Shopping/PMax",
        "module": "预算倾斜",
        "sop_id": "S073",
        "sop_name": "Hero 商品预算倾斜",
        "goal": "优化",
        "automation_level": "L3 (Recommendation)",
        "kpi": "Product ROAS",
        "trigger": "商品级",
        "action": """1. 识别 Hero 商品 (ROAS > Target * 1.2)
2. 增加预算分配
3. 创建独立 Campaign

建议：Hero 商品预算占比提升至 50%
建议：单独 Campaign 管理高价值商品""",
        "priority": "P1",
        "frequency": "每周",
        "output": "预算倾斜建议"
    },
    {
        "stage": 3,
        "campaign_type": "Shopping/PMax",
        "module": "预算倾斜",
        "sop_id": "S074",
        "sop_name": "低效率商品预算削减",
        "goal": "优化",
        "automation_level": "L4 (Auto-fix)",
        "kpi": "Product ROAS",
        "trigger": "商品级",
        "action": """1. 识别低效商品
2. 自动削减预算
3. 转移预算至 Hero 商品

IF Product ROAS < Target * 0.7 THEN 削减预算 30%""",
        "priority": "P1",
        "frequency": "每周",
        "output": "预算削减确认"
    },

    # Stage 4 - 战略复盘 (S075-S084)
    {
        "stage": 4,
        "campaign_type": "All",
        "module": "A/B 测试",
        "sop_id": "S075",
        "sop_name": "广告文案 A/B 测试设计与执行",
        "goal": "测试",
        "automation_level": "L3 (Recommendation)",
        "kpi": "CTR, CVR",
        "trigger": "广告级",
        "action": """1. 设计测试假设
2. 创建变体广告
3. 监控统计显著性
4. 实施获胜变体

建议：每次只测试一个变量
建议：样本量至少 1000 展示""",
        "priority": "P2",
        "frequency": "每月",
        "output": "A/B 测试报告"
    },
    {
        "stage": 4,
        "campaign_type": "All",
        "module": "A/B 测试",
        "sop_id": "S076",
        "sop_name": "出价策略 A/B 测试",
        "goal": "测试",
        "automation_level": "L3 (Recommendation)",
        "kpi": "CPA, ROAS",
        "trigger": "Campaign 级",
        "action": """1. 选择测试出价策略
2. 创建实验 Campaign
3. 对比性能指标
4. 选择最优策略

建议：测试期至少 14 天
建议：使用草稿和实验功能""",
        "priority": "P2",
        "frequency": "每季度",
        "output": "出价策略测试报告"
    },
    {
        "stage": 4,
        "campaign_type": "Lead Gen",
        "module": "线索质量",
        "sop_id": "S077",
        "sop_name": "MQL/SQL 转化率分析",
        "goal": "分析",
        "automation_level": "L2 (Diagnostic)",
        "kpi": "MQL rate, SQL rate",
        "trigger": "Campaign 级",
        "action": """1. 追踪线索转化漏斗
2. 计算 MQL/SQL 率
3. 识别质量瓶颈
4. 优化获取策略

IF MQL Rate < 50% THEN 优化表单字段
IF SQL Rate < 30% THEN 加强销售跟进""",
        "priority": "P1",
        "frequency": "每月",
        "output": "线索质量分析报告"
    },
    {
        "stage": 4,
        "campaign_type": "Lead Gen",
        "module": "线索质量",
        "sop_id": "S078",
        "sop_name": "高价值线索 Campaign 优化",
        "goal": "优化",
        "automation_level": "L3 (Recommendation)",
        "kpi": "Deal size, Close rate",
        "trigger": "Campaign 级",
        "action": """1. 分析线索成交数据
2. 识别高价值 Campaign
3. 优化预算分配
4. 扩大成功模式

建议：高成交率 Campaign 增加 30% 预算
建议：复制成功 Campaign 设置""",
        "priority": "P1",
        "frequency": "每月",
        "output": "高价值 Campaign 优化建议"
    },
    {
        "stage": 4,
        "campaign_type": "Lead Gen",
        "module": "线索质量",
        "sop_id": "S079",
        "sop_name": "低质量线索来源排除",
        "goal": "清理",
        "automation_level": "L4 (Auto-fix)",
        "kpi": "Lead quality score",
        "trigger": "来源级",
        "action": """1. 评估各来源线索质量
2. 识别低质量来源
3. 排除或降低出价
4. 优化定向策略

IF Lead Quality Score < 3 THEN 降低出价 50%""",
        "priority": "P1",
        "frequency": "每月",
        "output": "低质量来源排除列表"
    },
    {
        "stage": 4,
        "campaign_type": "All",
        "module": "战略复盘",
        "sop_id": "S080",
        "sop_name": "90 天账户表现复盘",
        "goal": "复盘",
        "automation_level": "L2 (Diagnostic)",
        "kpi": "QoQ growth",
        "trigger": "账户级",
        "action": """1. 收集 90 天数据
2. 分析趋势变化
3. 识别成功因素
4. 制定下季度计划

分析维度：Spend, Conversion, CPA, ROAS 趋势""",
        "priority": "P2",
        "frequency": "每季度",
        "output": "90 天复盘报告"
    },
    {
        "stage": 4,
        "campaign_type": "All",
        "module": "战略复盘",
        "sop_id": "S081",
        "sop_name": "渠道组合效率分析",
        "goal": "分析",
        "automation_level": "L2 (Diagnostic)",
        "kpi": "Channel ROAS, Incrementality",
        "trigger": "渠道级",
        "action": """1. 分析各渠道表现
2. 评估渠道协同效应
3. 计算边际效率
4. 优化渠道组合

建议：Search 占比 40-50%
建议：PMax 占比 30-40%""",
        "priority": "P2",
        "frequency": "每季度",
        "output": "渠道效率分析报告"
    },
    {
        "stage": 4,
        "campaign_type": "All",
        "module": "战略复盘",
        "sop_id": "S082",
        "sop_name": "预算分配战略调整",
        "goal": "战略",
        "automation_level": "L3 (Recommendation)",
        "kpi": "Portfolio ROAS",
        "trigger": "账户级",
        "action": """1. 评估当前预算分配
2. 对比效率曲线
3. 建议战略调整
4. 制定增长计划

建议：向高效 Campaign 倾斜预算
建议：测试新渠道 10-15% 预算""",
        "priority": "P2",
        "frequency": "每季度",
        "output": "预算战略建议"
    },
    {
        "stage": 4,
        "campaign_type": "Shopping/PMax",
        "module": "Feed 测试",
        "sop_id": "S083",
        "sop_name": "商品 Feed 元素 A/B 测试",
        "goal": "测试",
        "automation_level": "L3 (Recommendation)",
        "kpi": "CTR by feed element",
        "trigger": "商品级",
        "action": """1. 选择测试商品
2. 创建标题/图片变体
3. 运行测试
4. 应用获胜元素

建议：测试标题包含品牌名 vs 不包含
建议：测试生活方式图片 vs 白底图""",
        "priority": "P2",
        "frequency": "每月",
        "output": "Feed 测试报告"
    },
    {
        "stage": 4,
        "campaign_type": "Shopping/PMax",
        "module": "分层测试",
        "sop_id": "S084",
        "sop_name": "ROAS 分层 Campaign 测试",
        "goal": "测试",
        "automation_level": "L3 (Recommendation)",
        "kpi": "ROAS by tier",
        "trigger": "商品级",
        "action": """1. 按 ROAS 分层商品
2. 创建独立 Campaign
3. 测试不同出价策略
4. 优化分层结构

High ROAS: 激进出价获取更多转化
Low ROAS: 保守出价控制花费""",
        "priority": "P1",
        "frequency": "每季度",
        "output": "分层测试报告"
    },

    # Daily Alert (S085-S094)
    {
        "stage": "Daily",
        "campaign_type": "All",
        "module": "表现监控",
        "sop_id": "S085",
        "sop_name": "表现异常实时预警",
        "goal": "监控",
        "automation_level": "L1 (Monitoring)",
        "kpi": "Spend, CTR, CVR change",
        "trigger": "Campaign 级",
        "action": """1. 监控关键指标变化
2. 对比历史数据
3. 识别异常波动
4. 发送实时告警

IF Spend Change > 40% THEN 发送告警
IF CTR Drop > 30% THEN 发送告警""",
        "priority": "P0",
        "frequency": "实时监控",
        "output": "异常告警"
    },
    {
        "stage": "Daily",
        "campaign_type": "All",
        "module": "预算监控",
        "sop_id": "S086",
        "sop_name": "每日预算进度监控",
        "goal": "监控",
        "automation_level": "L1 (Monitoring)",
        "kpi": "Budget utilization",
        "trigger": "Campaign 级",
        "action": """1. 计算预算使用率
2. 预测耗尽时间
3. 监控消耗速度
4. 发送预警

IF Utilization > 90% THEN 发送预算告警
IF Pacing > 120% THEN 发送消耗过快告警""",
        "priority": "P1",
        "frequency": "每天",
        "output": "预算状态报告"
    },
    {
        "stage": "Daily",
        "campaign_type": "All",
        "module": "转化监控",
        "sop_id": "S087",
        "sop_name": "全渠道转化异常监控",
        "goal": "监控",
        "automation_level": "L1 (Monitoring)",
        "kpi": "Conversions, CPA",
        "trigger": "Campaign 级",
        "action": """1. 监控转化数量
2. 分析 CPA 变化
3. 识别异常模式
4. 发送告警

IF Conversions Drop > 50% THEN 发送转化告警
IF CPA Increase > 50% THEN 发送 CPA 告警""",
        "priority": "P0",
        "frequency": "每天",
        "output": "转化异常告警"
    },
    {
        "stage": "Daily",
        "campaign_type": "All",
        "module": "投放监控",
        "sop_id": "S088",
        "sop_name": "全渠道投放异常监测",
        "goal": "监控",
        "automation_level": "L1 (Monitoring)",
        "kpi": "Impressions, Status",
        "trigger": "Campaign 级",
        "action": """1. 监控展示量
2. 检查 Campaign 状态
3. 识别投放中断
4. 发送告警

IF Impressions = 0 AND Status = ENABLED THEN 发送投放告警""",
        "priority": "P0",
        "frequency": "每 4 小时",
        "output": "投放异常告警"
    },
    {
        "stage": "Daily",
        "campaign_type": "All",
        "module": "追踪监控",
        "sop_id": "S089",
        "sop_name": "全链路追踪健康巡检",
        "goal": "监控",
        "automation_level": "L1 (Monitoring)",
        "kpi": "Conversion tracking status",
        "trigger": "账户级",
        "action": """1. 检查转化追踪状态
2. 验证标签活性
3. 测试转化记录
4. 发送告警

IF No Conversions FOR 3 days AND Cost > $100 THEN 发送追踪告警""",
        "priority": "P0",
        "frequency": "每天",
        "output": "追踪健康报告"
    },
    {
        "stage": "Daily",
        "campaign_type": "Search",
        "module": "搜索词监控",
        "sop_id": "S090",
        "sop_name": "搜索词流量浪费预警",
        "goal": "监控",
        "automation_level": "L1 (Monitoring)",
        "kpi": "Search term cost",
        "trigger": "搜索词级",
        "action": """1. 监控搜索词花费
2. 识别高花费无转化词
3. 生成负向词建议
4. 发送预警

IF Search Term Cost > $30 AND Conversions = 0 THEN 发送浪费预警""",
        "priority": "P1",
        "frequency": "每天",
        "output": "搜索词浪费报告"
    },
    {
        "stage": "Daily",
        "campaign_type": "All",
        "module": "落地页监控",
        "sop_id": "S091",
        "sop_name": "落地页可用性与性能监控",
        "goal": "监控",
        "automation_level": "L1 (Monitoring)",
        "kpi": "LP status, Load time",
        "trigger": "页面级",
        "action": """1. 检测落地页状态
2. 监控加载速度
3. 检查移动端适配
4. 发送告警

IF HTTP Status != 200 THEN 发送故障告警
IF Load Time > 5s THEN 发送速度告警""",
        "priority": "P0",
        "frequency": "每小时",
        "output": "落地页状态报告"
    },
    {
        "stage": "Daily",
        "campaign_type": "All",
        "module": "政策监控",
        "sop_id": "S092",
        "sop_name": "广告政策合规实时监控",
        "goal": "监控",
        "automation_level": "L1 (Monitoring)",
        "kpi": "Approval status",
        "trigger": "广告级",
        "action": """1. 监控广告审核状态
2. 识别拒登广告
3. 分析拒登原因
4. 发送告警

IF Status = DISAPPROVED THEN 立即发送 P0 告警""",
        "priority": "P0",
        "frequency": "实时监控",
        "output": "政策告警"
    },
    {
        "stage": "Daily",
        "campaign_type": "All",
        "module": "政策监控",
        "sop_id": "S093",
        "sop_name": "受限投放状态监控",
        "goal": "监控",
        "automation_level": "L1 (Monitoring)",
        "kpi": "Serving status",
        "trigger": "广告级",
        "action": """1. 检查广告投放状态
2. 识别受限广告
3. 分析受限原因
4. 发送告警

IF Serving Status = LIMITED THEN 发送限制告警""",
        "priority": "P1",
        "frequency": "每天",
        "output": "受限状态报告"
    },
    {
        "stage": "Daily",
        "campaign_type": "All",
        "module": "政策监控",
        "sop_id": "S094",
        "sop_name": "申诉与修正追踪",
        "goal": "追踪",
        "automation_level": "L2 (Diagnostic)",
        "kpi": "Appeal status",
        "trigger": "广告级",
        "action": """1. 追踪申诉状态
2. 监控审核进度
3. 记录申诉结果
4. 更新策略

建议：被拒登后立即申诉
建议：记录拒登模式优化内容""",
        "priority": "P1",
        "frequency": "每天",
        "output": "申诉追踪报告"
    },
]

# 创建工作簿
wb = openpyxl.Workbook()

# 删除默认工作表
if 'Sheet' in wb.sheetnames:
    del wb['Sheet']

# 创建 SOP 工作表
ws_sop = wb.create_sheet('SOP')

# 定义表头
headers = [
    'SOP阶段', 'Campaign Type', '模块', 'SOP', '目标',
    '自动化程度', 'KPI / Signal', '触发层级', '执行动作 / SOP描述',
    '优先级', '频率', '产出物'
]

# 写入表头
for col_idx, header in enumerate(headers, 1):
    cell = ws_sop.cell(row=1, column=col_idx, value=header)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# 写入数据
row_idx = 2
for item in sop_data:
    ws_sop.cell(row=row_idx, column=1, value=f"Stage {item['stage']}")
    ws_sop.cell(row=row_idx, column=2, value=item['campaign_type'])
    ws_sop.cell(row=row_idx, column=3, value=item['module'])
    ws_sop.cell(row=row_idx, column=4, value=f"{item['sop_id']}: {item['sop_name']}")
    ws_sop.cell(row=row_idx, column=5, value=item['goal'])
    ws_sop.cell(row=row_idx, column=6, value=item['automation_level'])
    ws_sop.cell(row=row_idx, column=7, value=item['kpi'])
    ws_sop.cell(row=row_idx, column=8, value=item['trigger'])
    ws_sop.cell(row=row_idx, column=9, value=item['action'])
    ws_sop.cell(row=row_idx, column=10, value=item['priority'])
    ws_sop.cell(row=row_idx, column=11, value=item['frequency'])
    ws_sop.cell(row=row_idx, column=12, value=item['output'])

    # 设置行高和自动换行
    for col_idx in range(1, 13):
        cell = ws_sop.cell(row=row_idx, column=col_idx)
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    row_idx += 1

# 设置列宽
column_widths = [12, 15, 15, 40, 10, 15, 25, 12, 60, 10, 12, 25]
for col_idx, width in enumerate(column_widths, 1):
    ws_sop.column_dimensions[get_column_letter(col_idx)].width = width

# 创建 Daily Alert Setup 工作表
ws_daily = wb.create_sheet('Daily Alert Setup')

daily_headers = ['Alert Type', 'Threshold', 'Priority', 'Notification', 'Action']
for col_idx, header in enumerate(daily_headers, 1):
    cell = ws_daily.cell(row=1, column=col_idx, value=header)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="C0504D", end_color="C0504D", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")

daily_data = [
    ('Spend Increase > 40%', '40%', 'P1', 'Email + Slack', 'Check budget settings'),
    ('CTR Drop > 30%', '30%', 'P2', 'Email', 'Review ad relevance'),
    ('CPA Increase > 50%', '50%', 'P0', 'Email + SMS', 'Pause underperforming keywords'),
    ('Zero Conversions > 3 days', '3 days', 'P0', 'Email + Slack', 'Check tracking setup'),
    ('Budget Utilization > 90%', '90%', 'P1', 'Email', 'Increase or adjust budget'),
    ('Impressions = 0', '0', 'P0', 'Email + SMS', 'Check campaign status'),
    ('Ad Disapproved', 'DISAPPROVED', 'P0', 'Email + SMS', 'Fix and appeal immediately'),
    ('Landing Page Error', 'HTTP != 200', 'P0', 'Email + Slack', 'Fix LP or pause ads'),
    ('Search Term Waste > $30', '$30', 'P1', 'Email', 'Add negative keyword'),
    ('Tracking Inactive > 3 days', '3 days', 'P0', 'Email + Slack', 'Check GTM/code'),
]

for row_idx, data in enumerate(daily_data, 2):
    for col_idx, value in enumerate(data, 1):
        ws_daily.cell(row=row_idx, column=col_idx, value=value)
        ws_daily.cell(row=row_idx, column=col_idx).alignment = Alignment(vertical="top", wrap_text=True)

daily_widths = [30, 15, 12, 20, 40]
for col_idx, width in enumerate(daily_widths, 1):
    ws_daily.column_dimensions[get_column_letter(col_idx)].width = width

# 保存文件
output_path = 'Google AD SOP v2.xlsx'
wb.save(output_path)
print(f"Excel 文件已生成: {output_path}")
print(f"总 SOP 数量: {len(sop_data)}")

# 统计各阶段数量
stage_counts = {}
for item in sop_data:
    stage = item['stage']
    stage_counts[stage] = stage_counts.get(stage, 0) + 1

print("\n各阶段 SOP 分布:")
# 先输出数字阶段，再输出字符串阶段
for stage in [1, 2, 3, 4, 'Daily']:
    if stage in stage_counts:
        print(f"  Stage {stage}: {stage_counts[stage]} 个")
